from requests_oauthlib import OAuth2Session
from tkinter.filedialog import askopenfilename, asksaveasfilename
import openpyxl
from datetime import datetime
from time import perf_counter
from math import ceil
import webbrowser
import pickle
from rich import print_json
from os.path import splitext
import concurrent.futures
import json


client_id = r''
client_secret = r''
redirect_uri = 'https://service.projectplace.com/'
refresh_url = 'https://api.projectplace.com/oauth2/access_token'
comments = []




def SaveToken(token) :
    with open('token.pickle', 'wb') as file:
        pickle.dump(token, file)
def LoadToken() :
    with open('token.pickle', 'rb') as file:
        loaded_token = pickle.load(file)
    return loaded_token





def GetAuthorization() :
    oauth = OAuth2Session(client_id, redirect_uri=redirect_uri,scope="")
    authorization_url, state = oauth.authorization_url(
        'https://api.projectplace.com/oauth2/authorize',
        access_type="offline", prompt="select_account")
    print(f'Abre en el navegador {authorization_url}\nY autoriza el acceso.')
    webbrowser.open(authorization_url)
    authorization_response = input('Ingresa la URL a la que redirige el navegador despues de otorgar acceso:\n')
    print("Obteniendo token, porfavor espera.")
    token = oauth.fetch_token(
        'https://api.projectplace.com/oauth2/access_token',
        authorization_response=authorization_response,
        client_secret=client_secret)
    SaveToken(token)


    return oauth, authorization_url,state, token
def RefreshToken(token) :
    extra = {
        'client_id': client_id,
        'client_secret': client_secret,
    }

    oauth = OAuth2Session(client_id, token=token)
    token= oauth.refresh_token(refresh_url, **extra)
    return oauth, token
if __name__ == "__main__":
    try:
        start = perf_counter()
        try:
            token = LoadToken()
            oauth,token = RefreshToken(token=token)
        except FileNotFoundError:
            print("No se encontro el token, creando nuevo token")
            oauth, authorization_url,state, token=GetAuthorization()
            exit()
        SaveToken(token)
        print('Abriendo tabla de excel...')
        path = askopenfilename()
        try:
            workbook = openpyxl.load_workbook(path)
        except openpyxl.utils.exceptions.InvalidFileException as e_file:
            print("Ruta invalida")
            exit()
        print('Tabla de excel abierta')
        worksheet= workbook.active

        cID = worksheet.cell(column=61,row=2).value
        
        wsID = worksheet.cell(column=72,row=2).value

        maxRows = worksheet.max_row

        api_url = "https://api.projectplace.com/"
        def getData(x):
            cardID = worksheet.cell(column=61,row=2+x).value
            workspaceID = worksheet.cell(column=72,row=2+x).value
            list = []
            Card_Events = oauth.get(f"{api_url}1/feeds/{workspaceID}/history?hub=1&id={cardID}&type=action")
                # Check if response status == '200 OK'
            #print(f"Status: {Card_Events.status_code}, Text: {Card_Events.text}")
            if Card_Events.status_code == 200:
                Card_Events = Card_Events.json()
                #print(f"Card_Events len: {len(Card_Events)}")
                # Print to screen every event on the card 
                for Card_Event in Card_Events ["data"]:
                    if (Card_Event['action'] == 'change_status_action') :
                        list.append(
                            {
                            'cardID':cardID,
                            'card_event':Card_Event['action_verbose'],
                            'card_action':Card_Event['action'],
                            'user_name':Card_Event['user_verbose'],
                            'progress':Card_Event['progress_verbose'],
                            'ts': Card_Event['ts']
                            }
                        )
                    if (Card_Event['action'] == 'change_title_action') :
                        print_json(json.dumps(Card_Event) )
                        list.append(
                            {
                            'cardID':cardID,
                            'card_event':Card_Event['action_verbose'],
                            'card_action':Card_Event['action'],
                            'user_name':Card_Event['user_verbose'],
                            'progress':Card_Event['action_new_title'],
                            'ts': Card_Event['ts']
                            }
                        )
            return list
        

        queryWorkbook = openpyxl.Workbook()
        querySheet = queryWorkbook.active
        
        querySheet.cell(column=1,row=1).value = 'Card ID'
        querySheet.cell(column=2,row=1).value = 'Nombre de la accion'
        querySheet.cell(column=3,row=1).value = 'Tipo de accion'
        querySheet.cell(column=4,row=1).value = 'Realizado por'
        querySheet.cell(column=5,row=1).value = 'Progreso'
        querySheet.cell(column=6,row=1).value = 'Fecha'

        print('Obteniendo datos desde la API...')
        with concurrent.futures.ThreadPoolExecutor() as executor:
            responseGen = executor.map(getData,range(2,maxRows-1))
        totalList = list(responseGen)
        
        i=0
        for queryCollection in totalList:
            for action in queryCollection:
                querySheet.cell(column=1,row=2+i).value = action['cardID']
                querySheet.cell(column=2,row=2+i).value = action['card_event']
                querySheet.cell(column=3,row=2+i).value = action['card_action']
                querySheet.cell(column=4,row=2+i).value = action['user_name']
                querySheet.cell(column=5,row=2+i).value = action['progress/action_new_title']
                querySheet.cell(column=6,row=2+i).value = str(datetime.fromtimestamp(action['ts']))
                i+=1
        print(f"Acciones obtenidas: {i}")
        newSavePath = splitext(path)
        queryWorkbook.save(newSavePath[0]+" - queryActionEvents"+newSavePath[1])
        
        stop = perf_counter()
        print("tiempo:",stop - start)
        exit()
        


    except KeyboardInterrupt:
        print("Programa terminado por Usuario (Keyboard Interruption)")
        exit()
    