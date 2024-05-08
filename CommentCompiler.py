from requests_oauthlib import OAuth2Session
from tkinter.filedialog import askopenfilename, asksaveasfilename
import openpyxl
from datetime import datetime
from time import perf_counter
import webbrowser
import pickle
from rich import print_json
from os.path import splitext

import concurrent.futures

import json

"""Ingresa entre las comillas los codigos client_secret y client ID obtenidos desde la página de ProjectPlace"""


client_id = r''
client_secret = r''
redirect_uri = 'https://service.projectplace.com/'



refresh_url = "https://api.projectplace.com/oauth2/access_token"
api_url = "https://api.projectplace.com/"


def SaveToken(token) :
    with open('token.pickle', 'wb') as file:
        pickle.dump(token, file)
def LoadToken() :
    with open('token.pickle', 'rb') as file:
        loaded_token = pickle.load(file)
    return loaded_token





def GetAuthorization() :
    oauth = OAuth2Session(client_id, redirect_uri=redirect_uri,scope="")
    authorization_url, state = oauth.authorization_url(
        'https://api.projectplace.com/oauth2/authorize',
        access_type="offline", prompt="select_account")
    print(f'Abre en el navegador {authorization_url}\nY autoriza el acceso.')
    webbrowser.open(authorization_url)
    authorization_response = input('Ingresa la URL a la que redirige el navegador despues de otorgar acceso:\n')
    print("Obteniendo token, porfavor espera.")
    token = oauth.fetch_token(
        'https://api.projectplace.com/oauth2/access_token',
        authorization_response=authorization_response,
        client_secret=client_secret)
    SaveToken(token)


    return oauth,state, token

def RefreshToken(token) :
    extra = {
        'client_id': client_id,
        'client_secret': client_secret,
    }

    oauth = OAuth2Session(client_id, token=token)
    token= oauth.refresh_token(refresh_url, **extra)
    return oauth, token


if __name__ == "__main__":
    start = perf_counter()
    try:
        token = LoadToken()
        oauth,token = RefreshToken(token=token)
    except FileNotFoundError:
        print("No se encontro el token, creando nuevo token")
        oauth,state, token=GetAuthorization()
        exit()
        
    SaveToken(token)



    print('Abriendo tabla de excel...')
    path = askopenfilename()
    try:
        workbook = openpyxl.load_workbook(path)
    except openpyxl.utils.exceptions.InvalidFileException as e_file:
        print("Ruta invalida")
        exit()
    print('Tabla de excel abierta')
    worksheet= workbook.active

    cID = worksheet.cell(column=61,row=2).value
    
    wsID = worksheet.cell(column=72,row=2).value

    maxRows = worksheet.max_row

    api_url = "https://api.projectplace.com/"
    def getData(x):
        cardID = worksheet.cell(column=61,row=2+x).value
        workspaceID = worksheet.cell(column=72,row=2+x).value
        list = []
        
        #print(f"Intentando hacer cardID:{cardID}")
        Card_Events = oauth.get(f"{api_url}3/conversations/comments/?count=100&item_id={cardID}&item_name=card&offset=0&sort_by=updated_at&sort_order=desc")
            # Check if response status == '200 OK'
        if Card_Events.status_code == 200:
            Card_Events = Card_Events.json()
            # Print to screen every event on the card 
            for Card_Comment in Card_Events ["data"]:
                list.append(
                        {
                        'item_id':Card_Comment['item_id'],
                        'created_at':Card_Comment['created_at'],
                        'createdBy':f"{Card_Comment['created_by']['first_name']} {Card_Comment['created_by']['last_name']}",
                        'text':Card_Comment['text']
                        }
                    )
        return list   
    
    
    queryWorkbook = openpyxl.Workbook()
    querySheet = queryWorkbook.active
    
    querySheet.cell(column=1,row=1).value = 'Card ID'
    querySheet.cell(column=2,row=1).value = 'Fecha de creacion'
    querySheet.cell(column=3,row=1).value = 'Creado por'
    querySheet.cell(column=4,row=1).value = 'Texto'
    
    
    
    x =range(2,maxRows-1)
    with concurrent.futures.ThreadPoolExecutor() as executor:
        responseGen = executor.map(getData,x)
    totalList = list(responseGen) 
    i = 0
    for collection in totalList:
        for comment in collection:
            querySheet.cell(column=1,row=2+i).value = comment['item_id']
            querySheet.cell(column=2,row=2+i).value = str(datetime.fromtimestamp(comment['created_at']))
            querySheet.cell(column=3,row=2+i).value = comment['createdBy']
            querySheet.cell(column=4,row=2+i).value = comment['text']
            i+=1
    print(f"Comentarios obtenidos: {i}")
    newSavePath = splitext(path)
    queryWorkbook.save(newSavePath[0]+" - queryComments"+newSavePath[1])
    
    stop = perf_counter()
    print("tiempo:",stop - start)
       
    exit()
    
    